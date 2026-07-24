"""Vendor Evaluation — per-contractor scored evaluations with editable grid.

Excel import → detect columns → user maps roles (weightage / actual / max / description) →
weighted score is auto-computed per row and rolled up to a Grand Total. Cells can also
contain spreadsheet-style formulas (`=SUM(A)`, `=A2*B2`, `=(A2/B2)*C2`, ...).

Roles:
    super_admin — full CRUD (create, import, edit, delete)
    admin       — create, import, edit; NO delete
    vendor_admin — VIEW ONLY, scoped to own contractor
"""
from __future__ import annotations

import io
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.db import db
from app.deps import get_current_user, require_roles
from app.helpers import audit
from app.schemas import (
    VendorEvalCreate, VendorEvalUpdate, VendorEvalMapping,
)
from app.utils import new_id, now_iso

router = APIRouter(prefix="/vendor-evaluations", tags=["vendor-evaluations"])


def _num(x) -> Optional[float]:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except Exception:
        return None


# ---------- Formula evaluator ----------
# Supports:
#   =SUM(A2:A10)          — range sum
#   =SUM(A)               — whole-column sum (data rows only)
#   =A2*B2, =(A2/B2)*C2   — basic arithmetic on cell refs
#   =AVG(A) / =MAX(A) / =MIN(A) — column aggregates
#
# Column letters follow Excel: A=col 0, B=col 1, ...
# Row numbers are 1-based over the flattened LEAF (non-section) rows list, with 1 being
# the first leaf row — matches what users see 1:1 with the grid.

def _col_index(letter: str, columns: list) -> Optional[int]:
    try:
        i = ord(letter.upper()) - ord("A")
        return i if 0 <= i < len(columns) else None
    except Exception:
        return None


def _cell_num(letter: str, row_num: int, flat_rows: list, columns: list) -> float:
    idx = _col_index(letter, columns)
    if idx is None or row_num < 1 or row_num > len(flat_rows):
        return 0.0
    val = flat_rows[row_num - 1].get("_computed", {}).get(columns[idx]["key"])
    if val is None:
        val = flat_rows[row_num - 1].get("cells", {}).get(columns[idx]["key"])
    n = _num(val)
    return n if n is not None else 0.0


def _eval_formula(expr: str, flat_rows: list, columns: list) -> Optional[float]:
    if not isinstance(expr, str) or not expr.startswith("="):
        return None
    e = expr[1:].strip()
    if not e:
        return None

    # 1. AGG(letter:letter) — range on same column
    def range_repl(m):
        c1, r1, c2, r2 = m.group(2), int(m.group(3)), m.group(4), int(m.group(5))
        i1 = _col_index(c1, columns)
        i2 = _col_index(c2, columns)
        if i1 is None or i1 != i2:
            return "0"
        key = columns[i1]["key"]
        vals = []
        for r in flat_rows[r1 - 1: r2]:
            v = _num(r.get("_computed", {}).get(key, r.get("cells", {}).get(key)))
            if v is not None:
                vals.append(v)
        fn = m.group(1).upper()
        if not vals:
            return "0"
        return str({"SUM": sum(vals), "AVG": sum(vals) / len(vals), "MAX": max(vals), "MIN": min(vals)}[fn])
    e = re.sub(r"(SUM|AVG|MAX|MIN)\(([A-Za-z])(\d+):([A-Za-z])(\d+)\)", range_repl, e)

    # 2. AGG(letter) — whole-column data
    def col_repl(m):
        c = m.group(2)
        i = _col_index(c, columns)
        if i is None:
            return "0"
        key = columns[i]["key"]
        vals = []
        for r in flat_rows:
            v = _num(r.get("_computed", {}).get(key, r.get("cells", {}).get(key)))
            if v is not None:
                vals.append(v)
        fn = m.group(1).upper()
        if not vals:
            return "0"
        return str({"SUM": sum(vals), "AVG": sum(vals) / len(vals), "MAX": max(vals), "MIN": min(vals)}[fn])
    e = re.sub(r"(SUM|AVG|MAX|MIN)\(([A-Za-z])\)", col_repl, e)

    # 3. Cell refs like A2, B3, …
    def cell_repl(m):
        return str(_cell_num(m.group(1), int(m.group(2)), flat_rows, columns))
    e = re.sub(r"([A-Za-z])(\d+)", cell_repl, e)

    # 4. Restricted eval — only arithmetic
    if not re.fullmatch(r"[0-9\.\s+\-*/()]+", e):
        return None
    try:
        return float(eval(e, {"__builtins__": {}}, {}))
    except Exception:
        return None


def _flatten_leaves(rows: list) -> list:
    """Return a flat list of leaf (non-section) rows in visual/creation order."""
    out = []
    def walk(rs):
        for r in rs or []:
            if r.get("is_section"):
                walk(r.get("sub_rows"))
            else:
                out.append(r)
    walk(rows)
    return out


def _row_weighted_score(cells: dict, mapping: dict) -> Optional[float]:
    if not mapping:
        return None
    w_col = mapping.get("weightage_col")
    a_col = mapping.get("actual_col")
    m_col = mapping.get("max_col")
    if not (w_col and a_col and m_col):
        return None
    w = _num(cells.get(w_col))
    a = _num(cells.get(a_col))
    mx = _num(cells.get(m_col))
    if w is None or a is None or mx is None or mx == 0:
        return None
    return (a / mx) * w


def _recompute(evaluation: dict) -> dict:
    """Walk rows (+ sub_rows) and compute: per-cell formulas, weighted_score per row,
    section rollups and grand_total."""
    mapping = evaluation.get("column_mapping") or {}
    columns = evaluation.get("columns") or []
    rows = evaluation.get("rows") or []

    # First pass: evaluate formulas top-down so later formulas can reference earlier computed cells
    leaves = _flatten_leaves(rows)
    for r in leaves:
        r["_computed"] = {}
    for r in leaves:
        for k, v in (r.get("cells") or {}).items():
            if isinstance(v, str) and v.startswith("="):
                computed = _eval_formula(v, leaves, columns)
                r["_computed"][k] = computed
    # Expose computed values back into cells for consumers (rows["cells"] stays as user-entered)
    # We store separately as `computed_cells` to preserve the raw formula in `cells`.
    for r in leaves:
        r["computed_cells"] = r.pop("_computed", {})

    # Second pass: weighted scores + rollup
    grand = 0.0
    grand_weight = 0.0

    def resolved(cells: dict, computed: dict, key: str):
        # Prefer computed value if present; else raw cell value
        if key and computed.get(key) is not None:
            return computed[key]
        return cells.get(key) if key else None

    def visit(row):
        nonlocal grand, grand_weight
        if row.get("is_section"):
            for c in (row.get("sub_rows") or []):
                visit(c)
            child_scores = [c.get("weighted_score") for c in (row.get("sub_rows") or []) if c.get("weighted_score") is not None]
            row["weighted_score"] = sum(child_scores) if child_scores else None
        else:
            cells = row.get("cells") or {}
            comp = row.get("computed_cells") or {}
            resolved_cells = {
                mapping.get("weightage_col"): resolved(cells, comp, mapping.get("weightage_col")),
                mapping.get("actual_col"): resolved(cells, comp, mapping.get("actual_col")),
                mapping.get("max_col"): resolved(cells, comp, mapping.get("max_col")),
            }
            ws = _row_weighted_score(resolved_cells, mapping)
            row["weighted_score"] = ws
            if ws is not None:
                grand += ws
                w = _num(resolved_cells.get(mapping.get("weightage_col")))
                if w is not None:
                    grand_weight += w

    for r in rows:
        visit(r)

    evaluation["grand_total"] = round(grand, 2) if grand else 0.0
    evaluation["total_weight"] = round(grand_weight, 2) if grand_weight else 0.0
    return evaluation


async def _scoped_filter(user: dict) -> dict:
    role = user["role"]
    if role in ("super_admin", "admin"):
        return {}
    if role == "vendor_admin":
        cid = user.get("contractor_id")
        return {"contractor_id": cid} if cid else {"_id": "__none__"}
    return {"_id": "__none__"}


@router.get("")
async def list_evaluations(user=Depends(require_roles("super_admin", "admin", "vendor_admin"))):
    f = await _scoped_filter(user)
    docs = await db.vendor_evaluations.find(f, {"_id": 0}).sort("updated_at", -1).to_list(500)
    # attach contractor name
    cids = list({d.get("contractor_id") for d in docs if d.get("contractor_id")})
    cmap = {c["id"]: c.get("name", "") for c in await db.contractors.find({"id": {"$in": cids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}
    out = []
    for d in docs:
        d["contractor_name"] = cmap.get(d.get("contractor_id"), "—")
        out.append({
            "id": d["id"],
            "title": d.get("title"),
            "period": d.get("period"),
            "region": d.get("region") or "",
            "contractor_id": d.get("contractor_id"),
            "contractor_name": d.get("contractor_name"),
            "grand_total": d.get("grand_total", 0.0),
            "total_weight": d.get("total_weight", 0.0),
            "row_count": len(d.get("rows") or []),
            "updated_at": d.get("updated_at"),
            "created_at": d.get("created_at"),
        })
    return out


@router.get("/{eid}")
async def get_evaluation(eid: str, user=Depends(require_roles("super_admin", "admin", "vendor_admin"))):
    f = await _scoped_filter(user)
    f["id"] = eid
    doc = await db.vendor_evaluations.find_one(f, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    if doc.get("contractor_id"):
        c = await db.contractors.find_one({"id": doc["contractor_id"]}, {"_id": 0, "name": 1})
        doc["contractor_name"] = (c or {}).get("name", "")
    return doc


@router.post("")
async def create_evaluation(payload: VendorEvalCreate, current=Depends(require_roles("super_admin", "admin"))):
    doc = {
        "id": new_id(),
        "title": payload.title,
        "period": payload.period or "",
        "contractor_id": payload.contractor_id,
        "region": payload.region or "",
        "columns": [],
        "rows": [],
        "column_mapping": {},
        "grand_total": 0.0,
        "total_weight": 0.0,
        "created_by": current["id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.vendor_evaluations.insert_one(doc)
    await audit(current, "vendor_eval.create", doc["id"], {"title": payload.title, "contractor_id": payload.contractor_id, "region": payload.region})
    return {k: v for k, v in doc.items() if k != "_id"}


@router.post("/{eid}/import")
async def import_excel(eid: str, file: UploadFile = File(...), current=Depends(require_roles("super_admin", "admin"))):
    ev = await db.vendor_evaluations.find_one({"id": eid})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse workbook: {e}")

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=400, detail="Workbook is empty")
    header_row = all_rows[0]
    columns = []
    for idx, cell in enumerate(header_row):
        label = (str(cell) if cell is not None else f"Column {idx + 1}").strip() or f"Column {idx + 1}"
        columns.append({"key": f"c{idx}", "label": label})

    rows = []
    for r in all_rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        cells = {}
        for i, val in enumerate(r):
            if i >= len(columns):
                break
            cells[columns[i]["key"]] = "" if val is None else (str(val) if not isinstance(val, (int, float)) else val)
        rows.append({
            "id": new_id(),
            "is_section": False,
            "label": "",
            "cells": cells,
            "sub_rows": [],
            "weighted_score": None,
        })

    ev["columns"] = columns
    ev["rows"] = rows
    ev["column_mapping"] = {}  # Reset mapping — user needs to remap after re-import
    ev = _recompute(ev)
    ev["updated_at"] = now_iso()
    await db.vendor_evaluations.update_one({"id": eid}, {"$set": {k: v for k, v in ev.items() if k != "_id"}})
    await audit(current, "vendor_eval.import", eid, {"columns": len(columns), "rows": len(rows), "file": file.filename})
    ev.pop("_id", None)
    return ev


@router.put("/{eid}/mapping")
async def update_mapping(eid: str, payload: VendorEvalMapping, current=Depends(require_roles("super_admin", "admin"))):
    ev = await db.vendor_evaluations.find_one({"id": eid})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    ev["column_mapping"] = payload.model_dump(exclude_none=True)
    ev = _recompute(ev)
    ev["updated_at"] = now_iso()
    await db.vendor_evaluations.update_one({"id": eid}, {"$set": {k: v for k, v in ev.items() if k != "_id"}})
    await audit(current, "vendor_eval.mapping", eid, ev["column_mapping"])
    ev.pop("_id", None)
    return ev


@router.put("/{eid}")
async def update_evaluation(eid: str, payload: VendorEvalUpdate, current=Depends(require_roles("super_admin", "admin"))):
    ev = await db.vendor_evaluations.find_one({"id": eid})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    body = payload.model_dump(exclude_unset=True)
    for k in ("title", "period", "contractor_id", "columns", "rows", "column_mapping"):
        if k in body:
            ev[k] = body[k]
    ev = _recompute(ev)
    ev["updated_at"] = now_iso()
    await db.vendor_evaluations.update_one({"id": eid}, {"$set": {k: v for k, v in ev.items() if k != "_id"}})
    await audit(current, "vendor_eval.update", eid, {"fields": list(body.keys())})
    ev.pop("_id", None)
    return ev


@router.get("/compare/scoreboard")
async def scoreboard(user=Depends(require_roles("super_admin", "admin", "vendor_admin"))):
    """Region-level + Vendor-level comparison for rewarding decisions.

    Returns:
        by_region: [{region, contractors: [{name, best_score, avg_score, latest_period, eval_count}], top: <winner>}]
        by_contractor: [{contractor_id, name, region, best_score, avg_score, latest_period, eval_count, evaluations: [{title, period, grand_total}]}]
        top_overall: <winning contractor>
    """
    f = await _scoped_filter(user)
    docs = await db.vendor_evaluations.find(f, {"_id": 0}).sort("updated_at", -1).to_list(2000)

    cids = list({d.get("contractor_id") for d in docs if d.get("contractor_id")})
    cmap = {c["id"]: c.get("name", "") for c in await db.contractors.find({"id": {"$in": cids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}

    per_contractor = {}
    for d in docs:
        cid = d.get("contractor_id") or "unknown"
        if cid not in per_contractor:
            per_contractor[cid] = {
                "contractor_id": cid,
                "name": cmap.get(cid, "—"),
                "region": d.get("region") or "—",
                "evaluations": [],
                "scores": [],
            }
        else:
            # If any doc has an explicit region, prefer it
            if d.get("region") and per_contractor[cid]["region"] in ("", "—"):
                per_contractor[cid]["region"] = d["region"]
        gt = float(d.get("grand_total") or 0)
        per_contractor[cid]["evaluations"].append({
            "id": d.get("id"), "title": d.get("title"), "period": d.get("period") or "",
            "grand_total": gt, "updated_at": d.get("updated_at"),
        })
        per_contractor[cid]["scores"].append(gt)

    contractors_out = []
    for cid, info in per_contractor.items():
        scores = info["scores"] or [0]
        info["best_score"] = round(max(scores), 2)
        info["avg_score"] = round(sum(scores) / len(scores), 2)
        info["latest_period"] = info["evaluations"][0]["period"] if info["evaluations"] else ""
        info["eval_count"] = len(info["evaluations"])
        del info["scores"]
        contractors_out.append(info)
    contractors_out.sort(key=lambda x: x["best_score"], reverse=True)

    # Group by region
    per_region = {}
    for c in contractors_out:
        region = c["region"] or "—"
        if region not in per_region:
            per_region[region] = {"region": region, "contractors": []}
        per_region[region]["contractors"].append({
            "contractor_id": c["contractor_id"], "name": c["name"],
            "best_score": c["best_score"], "avg_score": c["avg_score"],
            "latest_period": c["latest_period"], "eval_count": c["eval_count"],
        })
    regions_out = list(per_region.values())
    for r in regions_out:
        r["contractors"].sort(key=lambda x: x["best_score"], reverse=True)
        r["top"] = r["contractors"][0] if r["contractors"] else None
        r["region_best"] = r["contractors"][0]["best_score"] if r["contractors"] else 0
    regions_out.sort(key=lambda r: r["region_best"], reverse=True)

    top_overall = contractors_out[0] if contractors_out else None
    return {"by_region": regions_out, "by_contractor": contractors_out, "top_overall": top_overall}


@router.get("/{eid}/export.xlsx")
async def export_xlsx(eid: str, user=Depends(require_roles("super_admin", "admin", "vendor_admin"))):
    f = await _scoped_filter(user)
    f["id"] = eid
    ev = await db.vendor_evaluations.find_one(f, {"_id": 0})
    if not ev:
        raise HTTPException(status_code=404, detail="Evaluation not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    # Metadata
    ws["A1"] = ev.get("title", "")
    ws["A1"].font = Font(bold=True, size=14)
    contractor = await db.contractors.find_one({"id": ev.get("contractor_id")}, {"_id": 0, "name": 1})
    ws["A2"] = f"Contractor: {(contractor or {}).get('name', '')}  ·  Region: {ev.get('region') or '—'}  ·  Period: {ev.get('period') or '—'}"
    ws["A2"].font = Font(italic=True, color="666666")

    columns = ev.get("columns") or []
    header_row = 4
    ws.cell(row=header_row, column=1, value="Description").font = Font(bold=True)
    for i, c in enumerate(columns):
        ws.cell(row=header_row, column=i + 2, value=c.get("label", "")).font = Font(bold=True)
    ws.cell(row=header_row, column=len(columns) + 2, value="Weighted Score").font = Font(bold=True)
    header_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")
    for col in range(1, len(columns) + 3):
        ws.cell(row=header_row, column=col).fill = header_fill

    row_cursor = header_row + 1

    def write_row(r: dict, depth: int = 0):
        nonlocal row_cursor
        indent = "  " * depth
        label = (r.get("label") or "").strip()
        ws.cell(row=row_cursor, column=1, value=f"{indent}{label}")
        if r.get("is_section"):
            for c in range(1, len(columns) + 3):
                ws.cell(row=row_cursor, column=c).font = Font(bold=True)
                ws.cell(row=row_cursor, column=c).fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        cells = r.get("cells") or {}
        for i, c in enumerate(columns):
            v = cells.get(c["key"], "")
            # Store as number if possible for spreadsheet-native formulas
            n = _num(v)
            ws.cell(row=row_cursor, column=i + 2, value=n if n is not None else v)
        ws.cell(row=row_cursor, column=len(columns) + 2, value=r.get("weighted_score"))
        ws.cell(row=row_cursor, column=len(columns) + 2).alignment = Alignment(horizontal="right")
        row_cursor += 1
        for sub in r.get("sub_rows") or []:
            write_row(sub, depth + 1)

    for r in ev.get("rows") or []:
        write_row(r)

    # Grand total
    total_row = row_cursor + 1
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=len(columns) + 2, value=ev.get("grand_total") or 0).font = Font(bold=True)
    ws.cell(row=total_row, column=len(columns) + 2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=len(columns) + 2).fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    ws.cell(row=total_row, column=len(columns) + 2).font = Font(bold=True, color="FFFFFF")

    # Autofit-ish widths
    ws.column_dimensions["A"].width = 40
    for i in range(len(columns) + 1):
        ws.column_dimensions[chr(ord("B") + i)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = re.sub(r"[^A-Za-z0-9_.-]+", "_", (ev.get("title") or "evaluation")) + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/{eid}")
async def delete_evaluation(eid: str, current=Depends(require_roles("super_admin"))):
    res = await db.vendor_evaluations.delete_one({"id": eid})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Evaluation not found")
    await audit(current, "vendor_eval.delete", eid)
    return {"ok": True}
